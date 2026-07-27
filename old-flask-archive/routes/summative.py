"""
Summative Competence Assessments — TVET CDACC grading.
Enter competence per unit per trainee
(Mastery / Proficient / Competent / Not Yet Competent / CRNM),
analyse unit performance, download reports, and generate graduation lists.
"""

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, abort, make_response,
)
from functools import wraps
from datetime import datetime
from io import BytesIO
from db import get_service_client
from auth_utils import current_user, login_required
from report_utils import (
    excel_letterhead, style_header_row, excel_signature_block,
    pdf_letterhead, pdf_header_style_cmds, pdf_signature_block,
)

summative_bp = Blueprint("summative", __name__, url_prefix="/summative")

# TVET CDACC competency grading:
#   M (80-100) · P (65-79) · C (50-64) · NYC (0-49) · CRNM (requirement not met)
COMPETENCE_LEVELS = (
    ("mastery", "Mastery"),
    ("proficient", "Proficient"),
    ("competent", "Competent"),
    ("not_yet_competent", "Not Yet Competent"),
    ("crnm", "Course Requirement Not Met"),
)
COMPETENCE_KEYS = {k for k, _ in COMPETENCE_LEVELS}
PASSING = frozenset({"mastery", "proficient", "competent"})
COMP_ABBR = {
    "mastery": "M",
    "proficient": "P",
    "competent": "C",
    "not_yet_competent": "NYC",
    "crnm": "CRNM",
}
COMP_LABEL = dict(COMPETENCE_LEVELS)

PORTAL_BASE = {
    "trainer":     "trainer/base.html",
    "dept_admin":  "dept_admin/base.html",
    "super_admin": "super_admin/base.html",
}

INSTITUTE = "THIKA TECHNICAL TRAINING INSTITUTE"
NAVY = "0F2744"
GOLD = "B45309"


def _portal(role: str) -> str:
    return PORTAL_BASE.get(role, "trainer/base.html")


def _is_admin(user) -> bool:
    return user.get("role") in ("dept_admin", "super_admin")


def _template_ctx(user):
    admin = _is_admin(user)
    return {
        "portal_base": _portal(user.get("role")),
        "is_admin": admin,
        # Trainers may view summative results; only HOD / Super Admin enter competence
        "can_enter_competence": admin,
        "show_trainer_filter": admin,
        "competence_levels": COMPETENCE_LEVELS,
    }


def staff_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        user = current_user()
        if not user or user.get("role") not in ("trainer", "dept_admin", "super_admin"):
            flash("Access denied.", "error")
            abort(403)
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        user = current_user()
        if not user or user.get("role") not in ("dept_admin", "super_admin"):
            flash("Access denied. Graduation lists are for admins.", "error")
            abort(403)
        return f(*args, **kwargs)
    return decorated


def _parse_year_term(default_year=None):
    year = request.args.get("year", default_year or datetime.now().year, type=int)
    term_raw = request.args.get("term", "")
    term = None
    if term_raw not in ("", None):
        try:
            term = int(term_raw)
        except (TypeError, ValueError):
            term = None
    return year, term


def _normalize_competence(comp):
    """Map legacy values onto the CDACC scale."""
    if comp == "exempt":
        return "not_yet_competent"
    if comp == "fail":
        return "crnm"
    return comp


def _trainer_class_units(db, user):
    return (db.table("class_units")
              .select("class_id, unit_id, trainer_id, units(id,code,name), classes(id,name)")
              .eq("trainer_id", user["id"])
              .execute().data or [])


def _dept_class_units(db, dept_id):
    classes = (db.table("classes")
                 .select("id, name")
                 .eq("department_id", dept_id)
                 .order("name")
                 .execute().data or [])
    class_ids = [c["id"] for c in classes]
    if not class_ids:
        return [], classes
    cu = (db.table("class_units")
            .select("class_id, unit_id, trainer_id, units(id,code,name), classes(id,name)")
            .in_("class_id", class_ids)
            .execute().data or [])
    return cu, classes


def _all_class_units(db):
    classes = (db.table("classes")
                 .select("id, name, department_id")
                 .order("name")
                 .execute().data or [])
    cu = (db.table("class_units")
            .select("class_id, unit_id, trainer_id, units(id,code,name), classes(id,name)")
            .execute().data or [])
    return cu, classes


def _scope_data(db, user):
    role = user.get("role")
    if role == "super_admin":
        return _all_class_units(db)
    if role == "dept_admin":
        dept_id = user.get("department_id")
        if not dept_id:
            return [], []
        return _dept_class_units(db, dept_id)
    cu = _trainer_class_units(db, user)
    class_map = {}
    for r in cu:
        c = r.get("classes") or {}
        if c.get("id"):
            class_map[c["id"]] = c["name"]
    classes = sorted([{"id": k, "name": v} for k, v in class_map.items()],
                     key=lambda x: x["name"])
    return cu, classes


def _detailed_class_units(db, user, class_id=None, trainer_id=None):
    """Class-unit rows with trainer names, scoped by role and optional filters."""
    sel = (
        "class_id, unit_id, trainer_id, "
        "units(id,code,name), classes(id,name,departments(name)), "
        "user_profiles!class_units_trainer_id_fkey(id, full_name)"
    )
    q = db.table("class_units").select(sel)
    role = user.get("role")
    if role == "trainer":
        q = q.eq("trainer_id", user["id"])
    elif role == "dept_admin":
        dept_id = user.get("department_id")
        if not dept_id:
            return []
        class_ids = [
            c["id"] for c in (
                db.table("classes").select("id").eq("department_id", dept_id).execute().data or []
            )
        ]
        if not class_ids:
            return []
        q = q.in_("class_id", class_ids)
    if class_id:
        q = q.eq("class_id", class_id)
    if trainer_id:
        q = q.eq("trainer_id", trainer_id)
    rows = q.execute().data or []
    rows.sort(key=lambda r: (
        (r.get("classes") or {}).get("name", ""),
        (r.get("units") or {}).get("code", ""),
    ))
    return rows


def _trainers_in_scope(db, user, class_id=None):
    rows = _detailed_class_units(db, user, class_id=class_id)
    seen = {}
    for r in rows:
        prof = r.get("user_profiles") or {}
        tid = r.get("trainer_id") or prof.get("id")
        if tid and tid not in seen:
            seen[tid] = {"id": tid, "full_name": prof.get("full_name") or "Trainer"}
    return sorted(seen.values(), key=lambda x: x["full_name"])


def _units_for_class(cu_rows, class_id, trainer_id=None):
    seen = {}
    for r in cu_rows:
        cid = (r.get("classes") or {}).get("id") or r.get("class_id")
        if cid != class_id:
            continue
        if trainer_id and r.get("trainer_id") != trainer_id:
            continue
        u = r.get("units") or {}
        uid = u.get("id") or r.get("unit_id")
        if uid and uid not in seen:
            seen[uid] = {
                "id": uid,
                "code": u.get("code", ""),
                "name": u.get("name", ""),
            }
    return sorted(seen.values(), key=lambda x: (x["code"] or x["name"]))


def _competence_query(db, class_id, unit_id=None, year=None, term=None):
    q = (db.table("summative_competences")
           .select("student_id, unit_id, competence, remarks, assessment_date, year, term, assessed_by")
           .eq("class_id", class_id))
    if unit_id:
        q = q.eq("unit_id", unit_id)
    if year:
        q = q.eq("year", year)
    if term:
        q = q.eq("term", term)
    return q.execute().data or []


def _enrolled_students(db, class_id):
    raw = (db.table("enrollments")
             .select("student_id, user_profiles(id, full_name, admission_no)")
             .eq("class_id", class_id)
             .execute().data or [])
    return sorted(raw, key=lambda s: (s.get("user_profiles") or {}).get("full_name", ""))


def _hub_stats(db, user, year):
    cu_rows, classes = _scope_data(db, user)
    class_ids = {c["id"] for c in classes}
    unit_keys = set()
    for r in cu_rows:
        cid = (r.get("classes") or {}).get("id") or r.get("class_id")
        uid = (r.get("units") or {}).get("id") or r.get("unit_id")
        if cid and uid:
            unit_keys.add((cid, uid))

    stats = {
        "total_assessed": 0, "passing": 0, "nyc": 0, "crnm": 0,
        "classes_count": len(classes), "units_count": len(unit_keys),
    }
    if not class_ids:
        return stats

    q = db.table("summative_competences").select("competence, class_id, unit_id").in_("class_id", list(class_ids))
    if year:
        q = q.eq("year", year)
    for r in q.execute().data or []:
        cid, uid = r.get("class_id"), r.get("unit_id")
        if (cid, uid) not in unit_keys and user.get("role") == "trainer":
            continue
        comp = _normalize_competence(r.get("competence"))
        stats["total_assessed"] += 1
        if comp in PASSING:
            stats["passing"] += 1
        elif comp == "not_yet_competent":
            stats["nyc"] += 1
        elif comp == "crnm":
            stats["crnm"] += 1
    return stats


def _build_unit_analyses(db, user, class_id="", trainer_id="", year=None, term=None):
    rows = _detailed_class_units(db, user, class_id=class_id or None, trainer_id=trainer_id or None)
    analyses = []
    totals = {"passing": 0, "nyc": 0, "crnm": 0, "assessed": 0}
    pass_rates = []

    for r in rows:
        cid = r.get("class_id") or (r.get("classes") or {}).get("id")
        u = r.get("units") or {}
        uid = u.get("id") or r.get("unit_id")
        if not cid or not uid:
            continue

        students = _enrolled_students(db, cid)
        enrolled = len(students)
        comp_rows = _competence_query(db, cid, unit_id=uid, year=year, term=term)
        cmap = {k: 0 for k, _ in COMPETENCE_LEVELS}
        cmap["missing"] = 0
        assessed_sids = set()
        for cr in comp_rows:
            comp = _normalize_competence(cr.get("competence"))
            if comp in cmap:
                cmap[comp] += 1
            assessed_sids.add(cr.get("student_id"))
        cmap["missing"] = max(0, enrolled - len(assessed_sids))

        assessed = sum(cmap[k] for k, _ in COMPETENCE_LEVELS)
        passing = cmap["mastery"] + cmap["proficient"] + cmap["competent"]
        pass_rate = round(passing / enrolled * 100, 1) if enrolled else 0
        completion = round(assessed / enrolled * 100, 1) if enrolled else 0
        pass_rates.append(pass_rate)

        prof = r.get("user_profiles") or {}
        analyses.append({
            "class_id": cid,
            "class_name": (r.get("classes") or {}).get("name", ""),
            "unit_id": uid,
            "unit_code": u.get("code", ""),
            "unit_name": u.get("name", ""),
            "trainer_name": prof.get("full_name", ""),
            "enrolled": enrolled,
            "counts": {
                "mastery": cmap["mastery"],
                "proficient": cmap["proficient"],
                "competent": cmap["competent"],
                "not_yet_competent": cmap["not_yet_competent"],
                "crnm": cmap["crnm"],
                "missing": cmap["missing"],
            },
            "pass_rate": pass_rate,
            "completion_rate": completion,
        })
        totals["passing"] += passing
        totals["nyc"] += cmap["not_yet_competent"]
        totals["crnm"] += cmap["crnm"]
        totals["assessed"] += assessed

    totals["avg_pass_rate"] = round(sum(pass_rates) / len(pass_rates), 1) if pass_rates else 0
    return analyses, totals


def _class_units_for_export(db, class_id, trainer_id=None):
    cu = (db.table("class_units")
            .select("unit_id, trainer_id, units(id, code, name)")
            .eq("class_id", class_id)
            .execute().data or [])
    unit_map = {}
    for r in cu:
        if trainer_id and r.get("trainer_id") != trainer_id:
            continue
        u = r.get("units") or {}
        uid = u.get("id") or r.get("unit_id")
        if uid:
            unit_map[uid] = {"id": uid, "code": u.get("code", ""), "name": u.get("name", "")}
    if not unit_map:
        cls = (db.table("classes").select("course_id").eq("id", class_id).limit(1).execute().data or [None])[0]
        if cls and cls.get("course_id"):
            for u in (db.table("units").select("id, code, name")
                        .eq("course_id", cls["course_id"]).order("code").execute().data or []):
                unit_map[u["id"]] = u
    return sorted(unit_map.values(), key=lambda x: (x.get("code") or x.get("name") or ""))


def _build_graduation_data(db, class_id, year=None, term=None, eligible_only=False):
    """Return (cls, meta, units, rows, stats) or raise ValueError."""
    cls = (db.table("classes")
             .select("id, name, course_id, department_id, departments(name), courses(name, code)")
             .eq("id", class_id)
             .limit(1)
             .execute().data or [None])[0]
    if not cls:
        raise ValueError("Class not found.")

    class_name = cls.get("name", "")
    dept_name = (cls.get("departments") or {}).get("name") or ""
    course_name = (cls.get("courses") or {}).get("name") or ""
    course_code = (cls.get("courses") or {}).get("code") or ""

    units = _class_units_for_export(db, class_id)
    students = _enrolled_students(db, class_id)

    q = (db.table("summative_competences")
           .select("student_id, unit_id, competence")
           .eq("class_id", class_id))
    if year:
        q = q.eq("year", year)
    if term:
        q = q.eq("term", term)
    competence_rows = q.execute().data or []

    cmap = {}
    for r in competence_rows:
        comp = _normalize_competence(r.get("competence"))
        cmap.setdefault(r["student_id"], {})[r["unit_id"]] = comp

    stats = {"eligible": 0, "not_eligible": 0, "total": 0, "pct_eligible": 0}
    rows = []
    for enr in students:
        profile = enr.get("user_profiles") or {}
        sid = enr.get("student_id") or profile.get("id")
        unit_results = {}
        all_met = True
        missing = nyc = crnm = 0
        for u in units:
            comp = (cmap.get(sid) or {}).get(u["id"])
            unit_results[u["id"]] = comp
            if not comp:
                all_met = False
                missing += 1
            elif comp not in PASSING:
                all_met = False
                if comp == "crnm":
                    crnm += 1
                else:
                    nyc += 1

        eligible = all_met and len(units) > 0
        if eligible:
            stats["eligible"] += 1
        else:
            stats["not_eligible"] += 1
        stats["total"] += 1

        rows.append({
            "student_id": sid,
            "full_name": profile.get("full_name", "—"),
            "admission_no": profile.get("admission_no", "—"),
            "unit_results": unit_results,
            "eligible": eligible,
            "missing": missing,
            "nyc": nyc,
            "crnm": crnm,
        })

    if eligible_only:
        rows = [r for r in rows if r["eligible"]]

    if stats["total"]:
        stats["pct_eligible"] = round(stats["eligible"] / stats["total"] * 100, 1)

    meta = {
        "class_name": class_name,
        "dept_name": dept_name,
        "course_name": course_name,
        "course_code": course_code,
    }
    return cls, meta, units, rows, stats


def _period_label(year, term):
    parts = []
    if year:
        parts.append(f"Year {year}")
    if term:
        parts.append(f"Term {term}")
    return " · ".join(parts) if parts else "All Periods"


def _excel_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="CBD5E1")
    return {
        "thin": thin,
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center"),
        "navy_fill": PatternFill("solid", fgColor=NAVY),
        "gold_fill": PatternFill("solid", fgColor="FEF3C7"),
        "ok_fill": PatternFill("solid", fgColor="DCFCE7"),
        "no_fill": PatternFill("solid", fgColor="FEE2E2"),
        "white": Font(bold=True, color="FFFFFF", size=10),
        "title_font": Font(bold=True, size=14, color=NAVY),
        "sub_font": Font(bold=True, size=11, color="334155"),
    }


def _attach_excel(resp, fname):
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


def _attach_pdf(resp, fname):
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ── Hub ───────────────────────────────────────────────────────────────────────

@summative_bp.route("/")
@staff_required
def hub():
    db = get_service_client()
    user = current_user()
    year, _ = _parse_year_term()
    cu_rows, _ = _scope_data(db, user)
    stats = _hub_stats(db, user, year)

    recent_units = []
    for r in cu_rows[:12]:
        c = r.get("classes") or {}
        u = r.get("units") or {}
        recent_units.append({
            "class_id": c.get("id") or r.get("class_id"),
            "class_name": c.get("name", ""),
            "unit_id": u.get("id") or r.get("unit_id"),
            "unit_code": u.get("code", ""),
            "unit_name": u.get("name", ""),
            "trainer_name": user.get("full_name") if user.get("role") == "trainer" else "",
        })

    if _is_admin(user):
        detailed = _detailed_class_units(db, user)[:12]
        recent_units = []
        for r in detailed:
            c = r.get("classes") or {}
            u = r.get("units") or {}
            prof = r.get("user_profiles") or {}
            recent_units.append({
                "class_id": c.get("id") or r.get("class_id"),
                "class_name": c.get("name", ""),
                "unit_id": u.get("id") or r.get("unit_id"),
                "unit_code": u.get("code", ""),
                "unit_name": u.get("name", ""),
                "trainer_name": prof.get("full_name", ""),
            })

    ctx = _template_ctx(user)
    ctx.update(stats=stats, recent_units=recent_units, year=year)
    return render_template("summative/hub.html", **ctx)


# ── Entry grid ────────────────────────────────────────────────────────────────

@summative_bp.route("/entry")
@staff_required
def entry():
    db = get_service_client()
    user = current_user()

    # Trainers are read-only for summative — competence entry is admin-only
    if not _is_admin(user):
        flash(
            "Summative competence entry is reserved for Department / Super Admins. "
            "You can view unit performance and download reports.",
            "warning",
        )
        return redirect(url_for("summative.analysis"))

    cu_rows, class_list = _scope_data(db, user)

    class_id = request.args.get("class_id", "")
    unit_id = request.args.get("unit_id", "")
    year, term = _parse_year_term()

    units_list = _units_for_class(cu_rows, class_id) if class_id else []
    students_list = _enrolled_students(db, class_id) if class_id else []
    competence_map = {}

    if class_id and unit_id:
        rows = _competence_query(db, class_id, unit_id=unit_id, year=year, term=term)
        for r in rows:
            r["competence"] = _normalize_competence(r.get("competence"))
            competence_map[r["student_id"]] = r

    summary = {k: 0 for k, _ in COMPETENCE_LEVELS}
    for r in competence_map.values():
        c = r.get("competence")
        if c in summary:
            summary[c] += 1

    ctx = _template_ctx(user)
    ctx.update(
        class_list=class_list,
        units_list=units_list,
        students_list=students_list,
        competence_map=competence_map,
        summary=summary,
        class_id=class_id,
        unit_id=unit_id,
        year=year,
        term=term or "",
        can_edit=True,
        active_tab="entry",
    )
    return render_template("summative/entry.html", **ctx)


@summative_bp.route("/save", methods=["POST"])
@staff_required
def save_competence():
    """AJAX — upsert one trainee competence for a unit (admin only)."""
    db = get_service_client()
    user = current_user()

    if not _is_admin(user):
        return jsonify({
            "success": False,
            "message": "Trainers cannot enter summative competence. Contact your HOD.",
        }), 403

    data = request.get_json() or {}

    student_id = (data.get("student_id") or "").strip()
    unit_id = (data.get("unit_id") or "").strip()
    class_id = (data.get("class_id") or "").strip()
    competence = (data.get("competence") or "").strip()
    remarks = (data.get("remarks") or "").strip() or None
    year = data.get("year") or datetime.now().year
    term = data.get("term") or None

    if not all([student_id, unit_id, class_id, competence]):
        return jsonify({"success": False, "message": "Missing required fields"}), 400
    if competence not in COMPETENCE_KEYS:
        return jsonify({"success": False, "message": "Invalid competence level"}), 400

    try:
        year = int(year)
    except (TypeError, ValueError):
        year = datetime.now().year
    try:
        term = int(term) if term not in (None, "", "null") else None
    except (TypeError, ValueError):
        term = None

    cls = (db.table("classes")
             .select("department_id")
             .eq("id", class_id)
             .limit(1)
             .execute().data or [None])[0]
    department_id = (cls or {}).get("department_id")

    payload = {
        "student_id": student_id,
        "unit_id": unit_id,
        "class_id": class_id,
        "department_id": department_id,
        "competence": competence,
        "assessed_by": user["id"],
        "assessment_date": datetime.now().date().isoformat(),
        "year": year,
        "term": term,
        "remarks": remarks,
    }

    try:
        existing = (db.table("summative_competences")
                      .select("id")
                      .eq("student_id", student_id)
                      .eq("unit_id", unit_id)
                      .eq("class_id", class_id)
                      .limit(1)
                      .execute().data or [])
        if existing:
            db.table("summative_competences").update(payload).eq("id", existing[0]["id"]).execute()
        else:
            db.table("summative_competences").insert(payload).execute()
        return jsonify({"success": True, "competence": competence})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ── Performance analysis ────────────────────────────────────────────────────────

@summative_bp.route("/analysis")
@staff_required
def analysis():
    db = get_service_client()
    user = current_user()
    _, class_list = _scope_data(db, user)

    class_id = request.args.get("class_id", "")
    trainer_id = request.args.get("trainer_id", "")
    year, term = _parse_year_term()

    if user.get("role") == "trainer":
        trainer_id = user["id"]

    unit_analyses, totals = _build_unit_analyses(
        db, user, class_id=class_id, trainer_id=trainer_id, year=year, term=term,
    )
    trainer_list = _trainers_in_scope(db, user, class_id=class_id or None)

    ctx = _template_ctx(user)
    ctx.update(
        class_list=class_list,
        class_id=class_id,
        trainer_id=trainer_id,
        trainer_list=trainer_list,
        year=year,
        term=term or "",
        unit_analyses=unit_analyses,
        totals=totals,
        active_tab="analysis",
    )
    return render_template("summative/analysis.html", **ctx)


# ── Reports & downloads ───────────────────────────────────────────────────────

@summative_bp.route("/reports")
@staff_required
def reports():
    db = get_service_client()
    user = current_user()
    cu_rows, class_list = _scope_data(db, user)

    class_id = request.args.get("class_id", "")
    unit_id = request.args.get("unit_id", "")
    trainer_id = request.args.get("trainer_id", "")
    year, term = _parse_year_term()

    units_list = _units_for_class(cu_rows, class_id, trainer_id or None) if class_id else []
    trainer_list = _trainers_in_scope(db, user, class_id=class_id or None)

    ctx = _template_ctx(user)
    ctx.update(
        class_list=class_list,
        units_list=units_list,
        class_id=class_id,
        unit_id=unit_id,
        trainer_id=trainer_id,
        trainer_list=trainer_list,
        year=year,
        term=term or "",
        active_tab="reports",
    )
    return render_template("summative/reports.html", **ctx)


@summative_bp.route("/export/unit.xlsx")
@staff_required
def export_unit_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    db = get_service_client()
    user = current_user()
    class_id = request.args.get("class_id", "")
    unit_id = request.args.get("unit_id", "")
    year, term = _parse_year_term()

    if not (class_id and unit_id):
        flash("Select class and unit.", "error")
        return redirect(url_for("summative.reports"))

    if user.get("role") == "trainer":
        assigned = (db.table("class_units")
                      .select("id")
                      .eq("trainer_id", user["id"])
                      .eq("class_id", class_id)
                      .eq("unit_id", unit_id)
                      .limit(1)
                      .execute().data or [])
        if not assigned:
            abort(403)

    cls = (db.table("classes")
             .select("name, department_id, departments(name)")
             .eq("id", class_id).limit(1).execute().data or [{}])[0]
    unit = (db.table("units").select("code, name").eq("id", unit_id).limit(1).execute().data or [{}])[0]
    trainer_name = user.get("full_name", "")
    cu = (db.table("class_units")
            .select("user_profiles!class_units_trainer_id_fkey(full_name)")
            .eq("class_id", class_id).eq("unit_id", unit_id).limit(1).execute().data or [])
    if cu:
        trainer_name = (cu[0].get("user_profiles") or {}).get("full_name") or trainer_name

    students = _enrolled_students(db, class_id)
    comp_rows = _competence_query(db, class_id, unit_id=unit_id, year=year, term=term)
    cmap = {r["student_id"]: r for r in comp_rows}

    st = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summative Unit"
    total_cols = 6
    dept_name = (cls.get("departments") or {}).get("name", "")

    header_row = excel_letterhead(
        ws, "Summative Competence — Unit Report", total_cols,
        dept_name=dept_name,
        meta_lines=(
            f"Class: {cls.get('name', '')}  |  Unit: {unit.get('code', '')} – {unit.get('name', '')}",
            f"Trainer: {trainer_name}  |  {_period_label(year, term)}  |  "
            f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ),
    )

    headers = ["#", "Admission No.", "Trainee Name", "Competence", "Remarks", "Assessed"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, total_cols)

    for i, enr in enumerate(students, 1):
        p = enr.get("user_profiles") or {}
        sid = enr.get("student_id") or p.get("id")
        rec = cmap.get(sid) or {}
        comp = _normalize_competence(rec.get("competence", ""))
        row_i = header_row + i
        vals = [
            i, p.get("admission_no", ""), p.get("full_name", ""),
            COMP_LABEL.get(comp, "—"), rec.get("remarks", "") or "",
            rec.get("assessment_date", "") or "—",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.border = st["border"]
            cell.alignment = st["left"] if col == 3 else st["center"]

    for col, w in zip("ABCDEF", [5, 16, 28, 18, 30, 14]):
        ws.column_dimensions[col].width = w

    excel_signature_block(ws, header_row + len(students) + 2,
                          officer_title="Head of Department")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = (unit.get("code") or "unit").replace(" ", "_")[:20]
    term_s = f"_T{term}" if term else ""
    fname = f"TTTI_Summative_{safe}{term_s}_{year}.xlsx"
    return _attach_excel(make_response(buf.getvalue()), fname)


@summative_bp.route("/export/class.xlsx")
@staff_required
def export_class_excel():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    db = get_service_client()
    user = current_user()
    class_id = request.args.get("class_id", "")
    trainer_id = request.args.get("trainer_id", "")
    year, term = _parse_year_term()

    if not class_id:
        flash("Select a class.", "error")
        return redirect(url_for("summative.reports"))

    units = _class_units_for_export(db, class_id, trainer_id or None)
    students = _enrolled_students(db, class_id)
    comp_rows = _competence_query(db, class_id, year=year, term=term)
    if trainer_id:
        unit_ids = {u["id"] for u in units}
        comp_rows = [r for r in comp_rows if r.get("unit_id") in unit_ids]

    cmap = {}
    for r in comp_rows:
        cmap.setdefault(r["student_id"], {})[r["unit_id"]] = _normalize_competence(r.get("competence"))

    cls = (db.table("classes")
             .select("name, departments(name), courses(name)")
             .eq("id", class_id).limit(1).execute().data or [{}])[0]
    st = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Matrix"
    total_cols = 3 + len(units)

    dept = (cls.get("departments") or {}).get("name", "")
    course = (cls.get("courses") or {}).get("name", "")
    header_row = excel_letterhead(
        ws, "Summative Competence — Class Matrix", total_cols,
        dept_name=dept,
        meta_lines=(
            f"Class: {cls.get('name', '')}  |  Course: {course}",
            f"{_period_label(year, term)}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ),
    )

    headers = ["#", "Admission No.", "Trainee Name"] + [
        (u.get("code") or u.get("name") or "U")[:16] for u in units
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, total_cols)

    for i, enr in enumerate(students, 1):
        p = enr.get("user_profiles") or {}
        sid = enr.get("student_id") or p.get("id")
        row_i = header_row + i
        vals = [i, p.get("admission_no", ""), p.get("full_name", "")]
        for u in units:
            vals.append(COMP_ABBR.get((cmap.get(sid) or {}).get(u["id"]), "—"))
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.border = st["border"]
            cell.alignment = st["left"] if col == 3 else st["center"]

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 28
    for i in range(len(units)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 11

    excel_signature_block(ws, header_row + len(students) + 2,
                          officer_title="Head of Department")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = (cls.get("name") or "class").replace(" ", "_")[:36]
    term_s = f"_T{term}" if term else ""
    fname = f"TTTI_Summative_Class_{safe}{term_s}_{year}.xlsx"
    return _attach_excel(make_response(buf.getvalue()), fname)


@summative_bp.route("/export/class.pdf")
@staff_required
def export_class_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    db = get_service_client()
    class_id = request.args.get("class_id", "")
    trainer_id = request.args.get("trainer_id", "")
    year, term = _parse_year_term()

    if not class_id:
        flash("Select a class.", "error")
        return redirect(url_for("summative.reports"))

    units = _class_units_for_export(db, class_id, trainer_id or None)
    students = _enrolled_students(db, class_id)
    comp_rows = _competence_query(db, class_id, year=year, term=term)
    if trainer_id:
        unit_ids = {u["id"] for u in units}
        comp_rows = [r for r in comp_rows if r.get("unit_id") in unit_ids]
    cmap = {}
    for r in comp_rows:
        cmap.setdefault(r["student_id"], {})[r["unit_id"]] = _normalize_competence(r.get("competence"))

    cls = (db.table("classes")
             .select("name, departments(name), courses(name)")
             .eq("id", class_id).limit(1).execute().data or [{}])[0]

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    cell_s = ParagraphStyle("c", parent=styles["Normal"], fontSize=7, leading=9)

    dept = (cls.get("departments") or {}).get("name", "")
    course = (cls.get("courses") or {}).get("name", "")
    story = pdf_letterhead(
        "Summative Competence — Class Matrix", doc.width,
        dept_name=dept,
        meta_lines=[
            f"Class: {cls.get('name', '')}" + (f"  |  Course: {course}" if course else ""),
            f"{_period_label(year, term)}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ],
    )

    headers = ["#", "Adm.", "Name"] + [
        Paragraph(f"<b>{(u.get('code') or 'U')[:8]}</b>", cell_s) for u in units
    ]
    data = [headers]
    for i, enr in enumerate(students, 1):
        p = enr.get("user_profiles") or {}
        sid = enr.get("student_id") or p.get("id")
        row = [str(i), p.get("admission_no", ""), Paragraph(p.get("full_name", ""), cell_s)]
        for u in units:
            row.append(COMP_ABBR.get((cmap.get(sid) or {}).get(u["id"]), "—"))
        data.append(row)

    unit_w = max(10 * mm, min(16 * mm, (200 * mm) / max(len(units), 1)))
    col_widths = [8 * mm, 20 * mm, 42 * mm] + [unit_w] * len(units)
    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]
    style_cmds.extend(pdf_header_style_cmds(0))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story += pdf_signature_block(doc.width, officer_title="Head of Department")
    doc.build(story)
    buf.seek(0)
    safe = (cls.get("name") or "class").replace(" ", "_")[:36]
    term_s = f"_T{term}" if term else ""
    fname = f"TTTI_Summative_Class_{safe}{term_s}_{year}.pdf"
    return _attach_pdf(make_response(buf.getvalue()), fname)


@summative_bp.route("/export/trainer.xlsx")
@staff_required
def export_trainer_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    db = get_service_client()
    user = current_user()
    trainer_id = request.args.get("trainer_id", "") or (
        user["id"] if user.get("role") == "trainer" else ""
    )
    class_id = request.args.get("class_id", "")
    year, term = _parse_year_term()

    if not trainer_id:
        flash("Select a trainer.", "error")
        return redirect(url_for("summative.reports"))

    if user.get("role") == "trainer" and trainer_id != user["id"]:
        abort(403)

    trainer = (db.table("user_profiles").select("full_name")
                 .eq("id", trainer_id).limit(1).execute().data or [{}])[0]
    rows = _detailed_class_units(db, user, class_id=class_id or None, trainer_id=trainer_id)

    st = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Trainer Summative"
    total_cols = 7

    dept_names = sorted({
        d for d in (
            ((cu.get("classes") or {}).get("departments") or {}).get("name", "")
            for cu in rows
        ) if d
    })
    dept_name = ", ".join(dept_names)

    header_row = excel_letterhead(
        ws, "Summative Report — Trainer", total_cols,
        dept_name=dept_name,
        meta_lines=(
            f"Trainer: {trainer.get('full_name', '')}",
            f"{_period_label(year, term)}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ),
    )

    headers = ["Class", "Unit Code", "Unit Name", "Trainee", "Admission No.", "Competence", "Remarks"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, total_cols)

    row_i = header_row
    for cu in rows:
        cid = cu.get("class_id") or (cu.get("classes") or {}).get("id")
        u = cu.get("units") or {}
        uid = u.get("id") or cu.get("unit_id")
        class_name = (cu.get("classes") or {}).get("name", "")
        students = _enrolled_students(db, cid)
        comp_rows = _competence_query(db, cid, unit_id=uid, year=year, term=term)
        cmap = {r["student_id"]: r for r in comp_rows}
        for enr in students:
            p = enr.get("user_profiles") or {}
            sid = enr.get("student_id") or p.get("id")
            rec = cmap.get(sid) or {}
            comp = _normalize_competence(rec.get("competence", ""))
            row_i += 1
            vals = [
                class_name, u.get("code", ""), u.get("name", ""),
                p.get("full_name", ""), p.get("admission_no", ""),
                COMP_LABEL.get(comp, "—"), rec.get("remarks", "") or "",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.border = st["border"]
                cell.alignment = st["left"]

    for col, w in zip("ABCDEFG", [22, 12, 24, 26, 14, 18, 24]):
        ws.column_dimensions[col].width = w

    excel_signature_block(ws, row_i + 2, officer_title="Head of Department")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = (trainer.get("full_name") or "trainer").replace(" ", "_")[:24]
    term_s = f"_T{term}" if term else ""
    fname = f"TTTI_Summative_Trainer_{safe}{term_s}_{year}.xlsx"
    return _attach_excel(make_response(buf.getvalue()), fname)


# ── Graduation list ───────────────────────────────────────────────────────────

@summative_bp.route("/graduation-list")
@admin_required
def graduation_list():
    db = get_service_client()
    user = current_user()
    _, class_list = _scope_data(db, user)

    class_id = request.args.get("class_id", "")
    year, term = _parse_year_term()
    eligible_only = request.args.get("eligible_only", "") == "1"

    units, rows = [], []
    meta = {"class_name": "", "dept_name": "", "course_name": "", "course_code": ""}
    stats = {"eligible": 0, "not_eligible": 0, "total": 0, "pct_eligible": 0}

    if class_id:
        try:
            _, meta, units, rows, stats = _build_graduation_data(
                db, class_id, year=year, term=term, eligible_only=eligible_only,
            )
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("summative.graduation_list"))

    ctx = _template_ctx(user)
    ctx.update(
        class_list=class_list,
        class_id=class_id,
        class_name=meta["class_name"],
        dept_name=meta["dept_name"],
        course_name=meta["course_name"],
        course_code=meta["course_code"],
        year=year,
        term=term or "",
        eligible_only=eligible_only,
        units=units,
        rows=rows,
        stats=stats,
        active_tab="graduation",
    )
    return render_template("summative/graduation_list.html", **ctx)


def _graduation_export_params():
    class_id = request.args.get("class_id", "")
    year, term = _parse_year_term()
    eligible_only = request.args.get("eligible_only", "") == "1"
    if not class_id:
        flash("Select a class first.", "error")
        return None, None, None, None
    return class_id, year, term, eligible_only


@summative_bp.route("/graduation-list/export.xlsx")
@admin_required
def graduation_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    db = get_service_client()
    class_id, year, term, eligible_only = _graduation_export_params()
    if not class_id:
        return redirect(url_for("summative.graduation_list"))

    try:
        _, meta, units, rows, stats = _build_graduation_data(
            db, class_id, year=year, term=term, eligible_only=eligible_only,
        )
    except ValueError as e:
        flash(str(e), "error")
        return redirect(url_for("summative.graduation_list"))

    st = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Graduation List" if not eligible_only else "Eligible Graduates"

    total_cols = 4 + len(units) + 1
    last = get_column_letter(total_cols)

    header_row = excel_letterhead(
        ws, "Official Graduation List — Summative Competence", total_cols,
        dept_name=meta["dept_name"],
        meta_lines=(
            f"Course: {meta['course_name'] or '—'} ({meta['course_code'] or '—'})  |  "
            f"Class: {meta['class_name']}  |  {_period_label(year, term)}",
            f"Eligible: {stats['eligible']} ({stats['pct_eligible']}%)   ·   "
            f"Not Eligible: {stats['not_eligible']}   ·   Total: {stats['total']}   ·   "
            f"Units Required: {len(units)}   ·   Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ),
    )

    headers = ["#", "Admission No.", "Trainee Name"] + [
        (u.get("code") or u.get("name") or "Unit")[:18] for u in units
    ] + ["Graduation Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, total_cols)
    ws.row_dimensions[header_row].height = 28

    for i, r in enumerate(rows, 1):
        row_i = header_row + i
        values = [i, r["admission_no"], r["full_name"]]
        for u in units:
            values.append(COMP_ABBR.get(r["unit_results"].get(u["id"]), "—"))
        values.append("ELIGIBLE" if r["eligible"] else "NOT ELIGIBLE")
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.border = st["border"]
            cell.alignment = st["center"] if col != 3 else st["left"]
            if col == total_cols:
                cell.fill = st["ok_fill"] if r["eligible"] else st["no_fill"]
                cell.font = Font(bold=True, size=9, color="166534" if r["eligible"] else "991B1B")

    legend_row = header_row + len(rows) + 2
    ws.merge_cells(f"A{legend_row}:{last}{legend_row}")
    ws[f"A{legend_row}"] = (
        "Legend (TVET CDACC): M = Mastery (80-100%)  |  P = Proficient (65-79%)  |  C = Competent (50-64%)  |  "
        "NYC = Not Yet Competent (0-49%)  |  CRNM = Course Requirement Not Met  |  — = Missing.  "
        "Eligible = all units rated Mastery, Proficient, or Competent."
    )
    ws[f"A{legend_row}"].font = Font(size=8, italic=True, color="64748B")
    ws[f"A{legend_row}"].fill = st["gold_fill"]

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 28
    for i in range(len(units)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12
    ws.column_dimensions[get_column_letter(total_cols)].width = 16

    excel_signature_block(ws, legend_row + 1,
                          officer_title="Head of Department",
                          extra_officers=("Examination Officer",))

    if not eligible_only and stats["eligible"] > 0:
        ws2 = wb.create_sheet("Eligible Only")
        eligible_rows = [r for r in rows if r["eligible"]]
        ws2.merge_cells("A1:D1")
        ws2["A1"] = f"ELIGIBLE GRADUATES — {meta['class_name']}"
        ws2["A1"].font = st["title_font"]
        ws2["A2"] = f"Count: {len(eligible_rows)}  |  {_period_label(year, term)}"
        for col, h in enumerate(["#", "Admission No.", "Trainee Name", "Status"], 1):
            ws2.cell(row=4, column=col, value=h)
        style_header_row(ws2, 4, 4)
        for i, r in enumerate(eligible_rows, 1):
            ws2.cell(row=4 + i, column=1, value=i).border = st["border"]
            ws2.cell(row=4 + i, column=2, value=r["admission_no"]).border = st["border"]
            ws2.cell(row=4 + i, column=3, value=r["full_name"]).border = st["border"]
            c = ws2.cell(row=4 + i, column=4, value="ELIGIBLE")
            c.fill = st["ok_fill"]
            c.border = st["border"]

    ws_auth = wb.create_sheet("Authentication")
    ws_auth["A1"] = INSTITUTE
    ws_auth["A1"].font = st["title_font"]
    ws_auth["A2"] = "Graduation List Authentication"
    ws_auth["A2"].font = st["sub_font"]
    ws_auth["A4"] = f"Class: {meta['class_name']}"
    ws_auth["A5"] = f"Period: {_period_label(year, term)}"
    auth_r = 7
    for officer in ("Head of Department", "Examination Officer", "Chief Principal"):
        ws_auth.cell(row=auth_r, column=1, value=f"{officer}:").font = Font(bold=True, size=10)
        ws_auth.cell(
            row=auth_r + 1, column=1,
            value="Name: _______________________________    Sign: ______________________    Date: ________________",
        )
        auth_r += 3
    ws_auth.cell(row=auth_r, column=1, value="Official Stamp:").font = Font(bold=True, size=10)
    ws_auth.cell(row=auth_r + 1, column=1, value="(Affix official institute stamp here)").font = Font(
        size=9, italic=True, color="64748B")
    ws_auth.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = (meta["class_name"] or "class").replace(" ", "_")[:40]
    term_s = f"_T{term}" if term else ""
    suffix = "_Eligible" if eligible_only else ""
    fname = f"TTTI_Graduation_List_{safe}{term_s}_{year}{suffix}.xlsx"
    return _attach_excel(make_response(buf.getvalue()), fname)


@summative_bp.route("/graduation-list/export.pdf")
@admin_required
def graduation_pdf():
    """Official landscape graduation list — print/save as PDF in browser."""
    db = get_service_client()
    class_id, year, term, eligible_only = _graduation_export_params()
    if not class_id:
        return redirect(url_for("summative.graduation_list"))

    try:
        _, meta, units, rows, stats = _build_graduation_data(
            db, class_id, year=year, term=term, eligible_only=eligible_only,
        )
    except ValueError as e:
        flash(str(e), "error")
        return redirect(url_for("summative.graduation_list"))

    return render_template(
        "summative/graduation_list_pdf.html",
        meta=meta,
        units=units,
        rows=rows,
        stats=stats,
        year=year,
        term=term,
        period_label=_period_label(year, term),
        eligible_only=eligible_only,
        generated=datetime.now().strftime("%d %b %Y"),
        show_print_controls=True,
    )


@summative_bp.route("/api/units/<class_id>")
@staff_required
def api_units(class_id):
    db = get_service_client()
    user = current_user()
    cu_rows, _ = _scope_data(db, user)
    trainer_id = request.args.get("trainer_id", "") or None
    units = _units_for_class(cu_rows, class_id, trainer_id)
    return jsonify(units)
