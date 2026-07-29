"""
routes/liaison_officer.py — Industrial Liaison Officer blueprint.
Manages attachment placements, approves processes, monitors placement progress,
coordinates supervisors and attachment records.
"""

from flask import Blueprint, render_template, request, flash, redirect, url_for, Response
from auth_utils import login_required, liaison_officer_required, current_user, write_audit_log
from db import get_service_client
from routes.attachment_helpers import (
    list_periods, get_open_period, is_student_eligible,
    upload_placement_document, placement_status_label,
    notify_liaison_officers, get_grading_config, compute_weighted_grade,
    score_to_cdacc, MENTOR_CRITERIA, week_bounds, section_max,
)
from datetime import datetime, date, timedelta
from report_utils import (excel_letterhead, style_header_row,
                          excel_signature_block, pdf_letterhead,
                          pdf_header_style_cmds, pdf_signature_block)

liaison_officer_bp = Blueprint("liaison_officer", __name__)
@liaison_officer_bp.route("/dashboard")
@login_required
@liaison_officer_required
def dashboard():
    db = get_service_client()
    user = current_user()
    stats = {}
    pending_attachments = []
    active_attachments = []
    recent_logbooks = []

    try:
        stats["total"]    = db.table("industrial_attachments").select("id", count="exact").execute().count or 0
        stats["pending"]  = db.table("industrial_attachments").select("id", count="exact").eq("status", "pending").execute().count or 0
        stats["active"]   = db.table("industrial_attachments").select("id", count="exact").eq("status", "active").execute().count or 0
        stats["approved"] = db.table("industrial_attachments").select("id", count="exact").eq("status", "approved").execute().count or 0
        stats["companies"]= db.table("companies").select("id", count="exact").execute().count or 0

        pending_attachments = (db.table("industrial_attachments")
            .select("*, user_profiles!industrial_attachments_student_id_fkey(full_name, admission_no, departments(name)), companies(name, address)")
            .eq("status", "pending")
            .order("created_at", desc=True)
            .limit(15)
            .execute().data or [])

        active_attachments = (db.table("industrial_attachments")
            .select("*, user_profiles!industrial_attachments_student_id_fkey(full_name, admission_no), companies(name, address)")
            .eq("status", "active")
            .order("start_date", desc=True)
            .limit(10)
            .execute().data or [])

        recent_logbooks = (db.table("digital_logbook")
            .select("*, user_profiles!digital_logbook_student_id_fkey(full_name, admission_no), industrial_attachments!digital_logbook_attachment_id_fkey(companies(name))")
            .order("log_date", desc=True)
            .limit(8)
            .execute().data or [])
    except Exception as e:
        flash(f"Error loading dashboard: {e}", "danger")

    return render_template("liaison_officer/dashboard.html",
                           stats=stats,
                           pending_attachments=pending_attachments,
                           active_attachments=active_attachments,
                           recent_logbooks=recent_logbooks,
                           current_month=datetime.now().strftime("%B %Y"))


@liaison_officer_bp.route("/attachments")
@login_required
@liaison_officer_required
def attachments():
    db = get_service_client()
    status_filter = request.args.get("status", "")
    query = (db.table("industrial_attachments")
               .select("*, user_profiles!industrial_attachments_student_id_fkey(full_name, admission_no, departments(name)), companies(name, address, email, city, county)")
               .order("created_at", desc=True)
               .limit(200))
    if status_filter:
        query = query.eq("status", status_filter)
    attachments_list = query.execute().data or []
    return render_template("liaison_officer/attachments.html",
                           attachments=attachments_list, status_filter=status_filter)


@liaison_officer_bp.route("/attachments/<att_id>")
@login_required
@liaison_officer_required
def placement_detail(att_id):
    db = get_service_client()
    att = (db.table("industrial_attachments")
           .select("*, user_profiles!industrial_attachments_student_id_fkey(full_name, admission_no, mobile_number, department_id, departments(name)), companies(*)")
           .eq("id", att_id)
           .limit(1)
           .execute().data or [])
    if not att:
        flash("Placement not found.", "warning")
        return redirect(url_for("liaison_officer.attachments"))
    record = att[0]
    trainer_name = ""
    if record.get("institute_trainer_id"):
        tr = (db.table("user_profiles").select("full_name").eq("id", record["institute_trainer_id"]).limit(1).execute().data or [])
        trainer_name = tr[0].get("full_name", "") if tr else ""
    record["_trainer_name"] = trainer_name

    trainers = (db.table("user_profiles")
                .select("id, full_name")
                .eq("role", "trainer")
                .order("full_name")
                .execute().data or [])
    departments = db.table("departments").select("id, name").order("name").execute().data or []

    return render_template(
        "liaison_officer/placement_detail.html",
        att=record,
        trainers=trainers,
        departments=departments,
        placement_status_label=placement_status_label,
    )


@liaison_officer_bp.route("/attachments/<att_id>/review", methods=["POST"])
@login_required
@liaison_officer_required
def review_placement(att_id):
    db = get_service_client()
    user = current_user()
    action = request.form.get("action", "")
    comments = (request.form.get("comments") or "").strip()

    if action not in ("approve", "reject", "needs_info"):
        flash("Invalid review action.", "warning")
        return redirect(url_for("liaison_officer.placement_detail", att_id=att_id))

    try:
        payload = {
            "liaison_review_comments": comments or None,
            "liaison_reviewed_by": user["id"],
            "liaison_reviewed_at": datetime.utcnow().isoformat(),
        }
        if action == "approve":
            payload.update({
                "placement_status": "verified",
                "acceptance_letter_status": "approved",
                "status": "approved",
                "approved_by": user["id"],
                "approved_at": datetime.utcnow().isoformat(),
            })
            msg = "Placement verified. Assign an institute trainer and activate when ready."
        elif action == "reject":
            payload.update({
                "placement_status": "rejected",
                "acceptance_letter_status": "rejected",
                "status": "rejected",
            })
            msg = "Placement rejected."
        else:
            payload.update({
                "placement_status": "needs_info",
                "status": "pending",
            })
            msg = "Trainee notified to provide more information."

        db.table("industrial_attachments").update(payload).eq("id", att_id).execute()
        write_audit_log(f"placement_{action}", target=f"attachment:{att_id}")
        flash(msg, "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    return redirect(url_for("liaison_officer.placement_detail", att_id=att_id))


@liaison_officer_bp.route("/attachments/<att_id>/assign", methods=["POST"])
@login_required
@liaison_officer_required
def assign_and_activate(att_id):
    db = get_service_client()
    user = current_user()
    trainer_id = (request.form.get("institute_trainer_id") or "").strip()
    department_id = (request.form.get("department_id") or "").strip()
    activate = request.form.get("activate") == "1"

    try:
        current = (db.table("industrial_attachments")
                     .select("id, status, placement_status")
                     .eq("id", att_id).limit(1).execute().data or [])
        if not current:
            flash("Placement not found.", "warning")
            return redirect(url_for("liaison_officer.attachments"))

        record = current[0]
        ps = record.get("placement_status") or ""
        if record.get("status") not in ("approved",) and ps != "verified":
            flash("Placement must be verified before assignment/activation.", "warning")
            return redirect(url_for("liaison_officer.placement_detail", att_id=att_id))

        payload = {}
        if trainer_id:
            payload["institute_trainer_id"] = trainer_id
        if department_id:
            payload["department_id"] = department_id
        if activate:
            payload["status"] = "active"
        if payload:
            db.table("industrial_attachments").update(payload).eq("id", att_id).execute()
            write_audit_log("assign_attachment", target=f"attachment:{att_id}", detail=payload)
            flash("Trainer assigned and attachment activated." if activate else "Assignment saved.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    return redirect(url_for("liaison_officer.placement_detail", att_id=att_id))


@liaison_officer_bp.route("/attachments/<att_id>/approve", methods=["POST"])
@login_required
@liaison_officer_required
def approve_attachment(att_id):
    db = get_service_client()
    user = current_user()
    new_status = request.form.get("status", "approved")
    if new_status not in ("active", "rejected", "completed"):
        flash("Invalid status.", "warning")
        return redirect(url_for("liaison_officer.attachments"))
    try:
        current = (db.table("industrial_attachments")
                     .select("id, status, acceptance_letter_status")
                     .eq("id", att_id)
                     .limit(1)
                     .execute().data or [])
        if not current:
            flash("Attachment record was not found.", "warning")
            return redirect(url_for("liaison_officer.attachments"))

        record = current[0]
        if new_status == "active":
            ps = record.get("placement_status") or "verified"
            if record.get("status") not in ("approved",) and ps not in ("verified",):
                flash("Placement must be verified by liaison before activation.", "warning")
                return redirect(url_for("liaison_officer.placement_detail", att_id=att_id))

        db.table("industrial_attachments").update({"status": new_status}).eq("id", att_id).execute()
        write_audit_log("update_attachment_status",
                        target=f"Attachment {att_id} set to {new_status}")
        flash(f"Attachment status updated to {new_status}.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    return redirect(url_for("liaison_officer.attachments"))


@liaison_officer_bp.route("/companies")
@login_required
@liaison_officer_required
def companies():
    db = get_service_client()
    records = db.table("companies").select("*").order("name").execute().data or []
    return render_template("liaison_officer/companies.html", companies=records)


@liaison_officer_bp.route("/logbooks")
@login_required
@liaison_officer_required
def logbooks():
    import os
    db = get_service_client()
    supabase_url  = os.environ.get("SUPABASE_URL", "").strip()
    status_filter = request.args.get("status", "")
    adm_filter    = request.args.get("admission_no", "").strip().upper()

    query = (db.table("digital_logbook")
               .select("id, student_id, log_date, entry_time, tasks_performed, "
                       "skills_applied, hours_worked, challenges_encountered, "
                       "achievements, mentor_approval_status, mentor_comments, "
                       "trainer_comments, evidence_urls, created_at, "
                       "student:user_profiles!digital_logbook_student_id_fkey"
                       "(full_name, admission_no), "
                       "attachment:industrial_attachments!digital_logbook_attachment_id_fkey"
                       "(companies(name))")
               .order("log_date", desc=True)
               .limit(500))

    if status_filter:
        query = query.eq("mentor_approval_status", status_filter)

    records = query.execute().data or []

    if adm_filter:
        records = [r for r in records
                   if adm_filter in (r.get("student") or {}).get("admission_no", "").upper()]

    for entry in records:
        ev_paths = entry.get("evidence_urls") or []
        entry["_evidence"] = [
            {
                "url":  f"{supabase_url}/storage/v1/object/public/assessment-evidence/{p}",
                "ext":  p.rsplit(".", 1)[-1].lower() if "." in p else "bin",
                "name": p.rsplit("/", 1)[-1],
            }
            for p in ev_paths if p
        ]

    return render_template("liaison_officer/logbooks.html",
                           logbooks=records, status_filter=status_filter,
                           adm_filter=adm_filter)


@liaison_officer_bp.route("/logbooks/<log_id>/review", methods=["POST"])
@login_required
@liaison_officer_required
def review_logbook(log_id):
    db  = get_service_client()
    user = current_user()
    action  = request.form.get("action", "")
    comment = (request.form.get("comment") or "").strip()

    if action not in ("approve", "reject"):
        flash("Invalid action.", "error")
        return redirect(url_for("liaison_officer.logbooks"))

    new_status = "approved" if action == "approve" else "rejected"
    payload = {
        "mentor_approval_status": new_status,
        "mentor_approved_by":     user["id"],
        "mentor_approved_at":     datetime.utcnow().isoformat(),
    }
    if comment:
        payload["mentor_comments"] = comment

    try:
        db.table("digital_logbook").update(payload).eq("id", log_id).execute()
        write_audit_log(f"{action}_logbook", target=f"Log {log_id}")
        flash(f"Log entry {new_status}.", "success")
    except Exception as exc:
        flash(f"Error: {exc}", "error")

    return redirect(url_for("liaison_officer.logbooks",
                            status=request.form.get("status_filter", "")))


# ── Attachment Export ──────────────────────────────────────────────────────────

def _get_period_range(year: int, period: str):
    ranges = {
        "1": (f"{year}-01-01", f"{year}-04-30"),
        "2": (f"{year}-05-01", f"{year}-08-31"),
        "3": (f"{year}-09-01", f"{year}-12-31"),
    }
    return ranges.get(period, (f"{year}-01-01", f"{year}-12-31"))


def _period_label(period: str) -> str:
    return {"1": "January–April", "2": "May–August", "3": "September–December"}.get(period, "Full Year")


@liaison_officer_bp.route("/attachments/export")
@login_required
@liaison_officer_required
def export_attachments():
    import io
    from datetime import date
    from flask import Response
    db = get_service_client()
    fmt    = request.args.get("format", "excel")
    period = request.args.get("period", "")
    year   = int(request.args.get("year", date.today().year))

    start_date, end_date = _get_period_range(year, period)
    rows_raw = (db.table("industrial_attachments")
                  .select("id, student_id, start_date, end_date, status, "
                          "student:user_profiles!industrial_attachments_student_id_fkey"
                          "(full_name, admission_no, mobile_number), "
                          "companies(name, address, contact_person, contact_phone)")
                  .gte("start_date", start_date)
                  .lte("start_date", end_date)
                  .order("student_id")
                  .execute().data or [])

    rows = []
    for r in rows_raw:
        st = r.get("student") or {}
        co = r.get("companies") or {}
        rows.append({
            "Admission No":       st.get("admission_no", ""),
            "Full Name":          st.get("full_name", ""),
            "Trainee Phone":      st.get("mobile_number", ""),
            "Company Attached":   co.get("name", ""),
            "Location / Address": co.get("address", ""),
            "Supervisor Name":    co.get("contact_person", ""),
            "Supervisor Phone":   co.get("contact_phone", ""),
            "Start Date":         r.get("start_date", ""),
            "End Date":           r.get("end_date", ""),
            "Status":             (r.get("status") or "").title(),
        })

    label = _period_label(period)
    title = "Industrial Attachments Register"
    headers = ["Admission No", "Full Name", "Trainee Phone", "Company Attached",
               "Location / Address", "Supervisor Name", "Supervisor Phone",
               "Start Date", "End Date", "Status"]
    user = current_user() or {}
    officer_name = user.get("full_name") or ""

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
        from report_utils import pdf_letterhead, pdf_header_style_cmds, pdf_signature_block

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=12*mm, bottomMargin=14*mm)
        avail = landscape(A4)[0] - 30 * mm
        col_headers = ["Adm No", "Full Name", "Phone", "Company", "Address",
                       "Supervisor", "Sup. Phone", "Start", "End", "Status"]
        keys = headers
        data = [col_headers] + [[r.get(k, "") for k in keys] for r in rows]
        col_widths_pt = [w * mm for w in [22, 48, 26, 48, 52, 40, 26, 20, 20, 18]]
        tbl = Table(data, colWidths=col_widths_pt, repeatRows=1)
        tbl.setStyle(TableStyle(list(pdf_header_style_cmds(0)) + [
            ("FONTSIZE",     (0, 0), (-1, 0), 8),
            ("FONTSIZE",     (0, 1), (-1, -1), 7.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("ALIGN",        (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ]))
        story = list(pdf_letterhead(
            doc_name=title,
            avail_width=avail,
            dept_name="Industrial Liaison Office — All Departments",
            meta_lines=(f"Period: {label} {year}", f"Total records: {len(rows)}"),
        ))
        story += [Spacer(1, 4), tbl]
        story += pdf_signature_block(
            avail,
            officer_title="Industrial Liaison Officer",
            officer_name=officer_name,
            extra_officers=("Head of Department",),
        )
        doc.build(story)
        buf.seek(0)
        fname = f"attachments_{label.replace('–','-')}_{year}.pdf"
        return Response(buf.read(), mimetype="application/pdf",
                        headers={"Content-Disposition": f"attachment; filename={fname}"})

    # Excel
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side
    from report_utils import (
        excel_letterhead, style_header_row, excel_signature_block, HEADER_BORDER_HEX,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Attachments"
    thin = Side(style="thin", color=HEADER_BORDER_HEX)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_cols = len(headers)

    next_row = excel_letterhead(
        ws,
        doc_name=title,
        total_cols=total_cols,
        dept_name="Industrial Liaison Office — All Departments",
        meta_lines=(f"Period: {label} {year}", f"Total records: {len(rows)}"),
    )
    hdr_row = next_row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=h)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    style_header_row(ws, hdr_row, total_cols)
    ws.row_dimensions[hdr_row].height = 28

    for row in rows:
        r = ws.max_row + 1
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col_idx, value=row.get(h, ""))
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col_idx in (2, 4, 5) else "center",
                                       vertical="center", wrap_text=True)

    from openpyxl.utils import get_column_letter
    for i, w in enumerate([16, 28, 18, 28, 32, 24, 18, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    excel_signature_block(
        ws, ws.max_row + 2,
        officer_title="Industrial Liaison Officer",
        officer_name=officer_name,
        extra_officers=("Head of Department",),
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"attachments_{label.replace('–','-')}_{year}.xlsx"
    return Response(buf.read(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Attachment periods (Step 1: liaison opens window) ─────────────────────────

@liaison_officer_bp.route("/periods", methods=["GET", "POST"])
@login_required
@liaison_officer_required
def attachment_periods():
    db = get_service_client()
    user = current_user()

    if request.method == "POST":
        action = request.form.get("action", "create")
        try:
            if action == "create":
                payload = {
                    "name": (request.form.get("name") or "").strip(),
                    "term": request.form.get("term"),
                    "year": int(request.form.get("year") or datetime.now().year),
                    "application_opens": request.form.get("application_opens"),
                    "application_closes": request.form.get("application_closes"),
                    "placement_deadline": request.form.get("placement_deadline") or None,
                    "notes": (request.form.get("notes") or "").strip() or None,
                    "is_open": request.form.get("is_open") == "1",
                    "created_by": user["id"],
                }
                intro = request.files.get("introduction_letter")
                if intro and intro.filename:
                    url, path = upload_placement_document(intro, user["id"], "intro_letter")
                    payload["introduction_letter_url"] = url
                    payload["introduction_letter_path"] = path
                db.table("attachment_periods").insert(payload).execute()
                flash("Attachment period created.", "success")

            elif action == "toggle":
                pid = request.form.get("period_id")
                is_open = request.form.get("is_open") == "1"
                db.table("attachment_periods").update({"is_open": is_open}).eq("id", pid).execute()
                flash("Period updated.", "success")

            elif action == "eligibility":
                pid = request.form.get("period_id")
                student_ids = request.form.getlist("student_ids")
                for sid in student_ids:
                    existing = (db.table("attachment_period_eligibility")
                                .select("id")
                                .eq("period_id", pid).eq("student_id", sid)
                                .limit(1).execute().data or [])
                    row = {
                        "period_id": pid,
                        "student_id": sid,
                        "is_eligible": True,
                        "introduction_letter_issued": request.form.get(f"intro_{sid}") == "1",
                        "approved_by": user["id"],
                        "approved_at": datetime.utcnow().isoformat(),
                    }
                    if existing:
                        db.table("attachment_period_eligibility").update(row).eq("id", existing[0]["id"]).execute()
                    else:
                        db.table("attachment_period_eligibility").insert(row).execute()
                flash(f"Marked {len(student_ids)} trainee(s) as eligible.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        return redirect(url_for("liaison_officer.attachment_periods"))

    periods = list_periods(db)
    students = (db.table("user_profiles")
                  .select("id, full_name, admission_no, departments(name)")
                  .eq("role", "student")
                  .order("full_name")
                  .limit(500)
                  .execute().data or [])

    # Load all classes (for class-based eligibility selection)
    classes_list = []
    try:
        classes_list = (db.table("classes")
                          .select("id, name, level, intake_year, intake_month, courses(name, code), departments(name)")
                          .order("name")
                          .execute().data or [])
    except Exception:
        pass

    # Build student_id → class info map from enrollments
    student_class_map = {}  # student_id → {"class_id": ..., "class_name": ...}
    try:
        enr_rows = (db.table("enrollments")
                      .select("student_id, class_id, classes(id, name, level, courses(code))")
                      .execute().data or [])
        for e in enr_rows:
            cls = e.get("classes") or {}
            student_class_map[e["student_id"]] = {
                "class_id":   cls.get("id", ""),
                "class_name": cls.get("name", ""),
                "level":      cls.get("level", ""),
                "course_code": (cls.get("courses") or {}).get("code", ""),
            }
    except Exception:
        pass

    # Build unique classes that have enrolled students (for class chips)
    seen_cids = set()
    classes_with_students = []
    class_student_counts = {}
    for s in students:
        sc = student_class_map.get(s["id"], {})
        cid = sc.get("class_id", "")
        if cid:
            class_student_counts[cid] = class_student_counts.get(cid, 0) + 1
            if cid not in seen_cids:
                seen_cids.add(cid)
                classes_with_students.append({
                    "id":          cid,
                    "name":        sc.get("class_name", ""),
                    "level":       sc.get("level", ""),
                    "course_code": sc.get("course_code", ""),
                })
    classes_with_students.sort(key=lambda x: x["name"])
    unclassified_count = sum(1 for s in students if not student_class_map.get(s["id"], {}).get("class_id"))

    eligibility = {}
    if periods and _table_ok_periods(db):
        try:
            rows = db.table("attachment_period_eligibility").select("*").execute().data or []
            for r in rows:
                eligibility.setdefault(r["period_id"], {})[r["student_id"]] = r
        except Exception:
            pass

    return render_template(
        "liaison_officer/periods.html",
        periods=periods,
        students=students,
        classes_list=classes_list,
        student_class_map=student_class_map,
        classes_with_students=classes_with_students,
        class_student_counts=class_student_counts,
        unclassified_count=unclassified_count,
        eligibility=eligibility,
        current_year=datetime.now().year,
        today=datetime.now().strftime("%Y-%m-%d"),
    )


def _table_ok_periods(db):
    try:
        db.table("attachment_periods").select("id").limit(1).execute()
        return True
    except Exception:
        return False


# ── Final grading ─────────────────────────────────────────────────────────────

@liaison_officer_bp.route("/attachments/<att_id>/grade", methods=["GET", "POST"])
@login_required
@liaison_officer_required
def grade_attachment(att_id):
    db = get_service_client()
    user = current_user()

    att_rows = (db.table("industrial_attachments")
                .select("*, user_profiles!industrial_attachments_student_id_fkey(full_name, admission_no, department_id)")
                .eq("id", att_id).limit(1).execute().data or [])
    if not att_rows:
        flash("Attachment not found.", "warning")
        return redirect(url_for("liaison_officer.attachments"))
    att = att_rows[0]
    dept_id = (att.get("user_profiles") or {}).get("department_id") or att.get("department_id")
    config = get_grading_config(db, dept_id)

    if request.method == "POST":
        try:
            weights = {
                "weight_gps_attendance": float(config.get("weight_gps_attendance", 10)),
                "weight_logbook": float(config.get("weight_logbook", 20)),
                "weight_mentor_eval": float(config.get("weight_mentor_eval", 30)),
                "weight_trainer_assessment": float(config.get("weight_trainer_assessment", 30)),
                "weight_final_report": float(config.get("weight_final_report", 10)),
            }
            scores = {
                "score_gps_attendance": float(request.form.get("score_gps_attendance") or 0),
                "score_logbook": float(request.form.get("score_logbook") or 0),
                "score_mentor_eval": float(request.form.get("score_mentor_eval") or 0),
                "score_trainer_assessment": float(request.form.get("score_trainer_assessment") or 0),
                "score_final_report": float(request.form.get("score_final_report") or 0),
            }
            for sk in ("score_gps_attendance", "score_logbook",
                       "score_trainer_assessment", "score_final_report"):
                scores[sk] = min(max(scores[sk], 0.0), section_max(weights, sk))

            mentor_fields = {}
            mentor_total = 0.0
            mentor_criteria_max = 0.0
            for field, _label, max_pts in MENTOR_CRITERIA:
                val = float(request.form.get(field) or 0)
                mentor_fields[field] = min(max(val, 0.0), float(max_pts))
                mentor_total += mentor_fields[field]
                mentor_criteria_max += float(max_pts)
            # Criteria sum to 100; scale into mentor section max marks (default /30)
            w_men = section_max(weights, "score_mentor_eval")
            if mentor_criteria_max > 0:
                scores["score_mentor_eval"] = round(
                    mentor_total * w_men / mentor_criteria_max, 2
                )
            else:
                scores["score_mentor_eval"] = 0.0
            scores["score_mentor_eval"] = min(
                max(scores["score_mentor_eval"], 0.0), w_men
            )

            weighted = compute_weighted_grade(scores, weights)
            grade = score_to_cdacc(weighted)

            grade_payload = {
                **scores,
                **mentor_fields,
                "weighted_total": weighted,
                "final_grade": grade,
                "graded_by": user["id"],
                "graded_at": datetime.utcnow().isoformat(),
            }
            existing = (db.table("attachment_grades")
                        .select("id").eq("attachment_id", att_id).limit(1).execute().data or [])
            if existing:
                db.table("attachment_grades").update(grade_payload).eq("attachment_id", att_id).execute()
            else:
                grade_payload["attachment_id"] = att_id
                db.table("attachment_grades").insert(grade_payload).execute()

            db.table("industrial_attachments").update({
                "final_grade": grade,
                "status": "completed",
            }).eq("id", att_id).execute()

            flash(f"Final grade recorded: {grade} ({weighted}%).", "success")
            return redirect(url_for("liaison_officer.placement_detail", att_id=att_id))
        except Exception as e:
            flash(f"Grading error: {e}", "danger")

    existing_grade = None
    try:
        rows = db.table("attachment_grades").select("*").eq("attachment_id", att_id).limit(1).execute().data or []
        existing_grade = rows[0] if rows else None
    except Exception:
        pass

    return render_template(
        "liaison_officer/grade_attachment.html",
        att=att,
        config=config,
        existing_grade=existing_grade,
        mentor_criteria=MENTOR_CRITERIA,
    )


# ── Industrial Attachment Marks (list + save) ─────────────────────────────────

from flask import jsonify as _jsonify

@liaison_officer_bp.route("/attachment-marks")
@login_required
@liaison_officer_required
def attachment_marks():
    db     = get_service_client()
    config = get_grading_config(db)

    raw = (db.table("industrial_attachments")
             .select("id, student_id, start_date, end_date, status, "
                     "companies(name), "
                     "student:user_profiles!industrial_attachments_student_id_fkey"
                     "(full_name, admission_no)")
             .order("created_at", desc=True)
             .limit(500)
             .execute().data or [])

    att_ids = [a["id"] for a in raw]
    grades_map = {}
    if att_ids:
        grade_rows = (db.table("attachment_grades")
                        .select("*")
                        .in_("attachment_id", att_ids)
                        .execute().data or [])
        grades_map = {g["attachment_id"]: g for g in grade_rows}

    for a in raw:
        a["_grade"] = grades_map.get(a["id"])

    total   = len(raw)
    graded  = sum(1 for a in raw if a.get("_grade"))
    pending = total - graded
    scores  = [float(a["_grade"]["weighted_total"]) for a in raw
               if a.get("_grade") and a["_grade"].get("weighted_total") is not None]
    avg     = round(sum(scores) / len(scores), 1) if scores else 0

    return render_template(
        "liaison_officer/attachment_marks.html",
        attachments=raw,
        config=config,
        stats={"total": total, "graded": graded, "pending": pending, "avg": avg},
    )


@liaison_officer_bp.route("/attachment-marks/<att_id>/save", methods=["POST"])
@login_required
@liaison_officer_required
def save_attachment_marks(att_id):
    db   = get_service_client()
    user = current_user()

    def _f(name):
        try:    return float(request.form.get(name) or 0)
        except: return 0.0

    scores = {
        "score_gps_attendance":    _f("score_gps_attendance"),
        "score_logbook":           _f("score_logbook"),
        "score_mentor_eval":       _f("score_mentor_eval"),
        "score_trainer_assessment": _f("score_trainer_assessment"),
        "score_final_report":      _f("score_final_report"),
    }
    config  = get_grading_config(db)
    weights = {k: v for k, v in config.items() if k.startswith("weight_")}
    scores  = {k: min(max(v, 0.0), section_max(weights, k)) for k, v in scores.items()}
    total   = compute_weighted_grade(scores, weights)
    grade   = score_to_cdacc(total)

    payload = {**scores, "weighted_total": total, "final_grade": grade,
               "graded_by": user["id"], "graded_at": datetime.utcnow().isoformat()}

    existing = (db.table("attachment_grades")
                  .select("id")
                  .eq("attachment_id", att_id)
                  .limit(1)
                  .execute().data or [])
    if existing:
        db.table("attachment_grades").update(payload).eq("attachment_id", att_id).execute()
    else:
        payload["attachment_id"] = att_id
        db.table("attachment_grades").insert(payload).execute()

    write_audit_log("save_attachment_marks", target=f"attachment:{att_id}",
                    detail={"grade": grade, "total": total})
    return _jsonify({"ok": True, "total": total, "grade": grade})


# ── Mentoring Tool / Logbook — Liaison Officer view ──────────────────────────

@liaison_officer_bp.route("/mentoring-tools")
@liaison_officer_required
def view_mentoring_tools():
    db = get_service_client()
    rows = (db.table("mentoring_tool_uploads")
              .select("*, student:user_profiles!student_id(full_name,admission_no,department_id), attachment:industrial_attachments!attachment_id(id,company_id,status)")
              .order("uploaded_at", desc=True)
              .execute().data or [])
    # attach company name
    co_ids = list({r["attachment"]["company_id"] for r in rows if r.get("attachment") and r["attachment"].get("company_id")})
    co_map = {}
    if co_ids:
        for co in (db.table("companies").select("id,name").in_("id", co_ids).execute().data or []):
            co_map[co["id"]] = co["name"]
    for r in rows:
        r["_company"] = co_map.get((r.get("attachment") or {}).get("company_id", ""), "—")
    return render_template("liaison_officer/mentoring_tools.html", uploads=rows)
